import io
from io import BytesIO

import pandas as pd
import streamlit as st
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from xhtml2pdf import pisa

st.set_page_config(layout="wide")
st.title("Conselho de Classe IFNMG Almenara")

st.info(
    "Envie um arquivo .docx com o boletim condensado."
)

# =========================================================
# 1. Extração dos dados do .docx (equivalente ao script antigo
#    convert_docx_to_csv.py, mas sem gravar CSV em disco)
# =========================================================

def extrair_dados_docx(arquivo_docx):
    """
    Recebe um arquivo .docx (em memória) e devolve um DataFrame
    com os dados do boletim, já filtrando alunos "CANCELADO".
    Se algo estiver errado, devolve (None, mensagem_de_erro).
    """
    doc = Document(arquivo_docx)
    tables = doc.tables

    if len(tables) < 2 or len(tables) % 2 != 0:
        return None, "Número inválido de tabelas encontrado no arquivo."

    boletins_filtrados = []
    cabecalho = []

    for i in range(0, len(tables), 2):
        tabela_dados = tables[i]
        tabela_boletim = tables[i + 1]

        nome_aluno = "Desconhecido"
        matricula_aluno = "Desconhecida"

        # Extração dos dados pessoais do aluno
        for row in tabela_dados.rows:
            cell_text = row.cells[0].text.strip()
            if ':' in cell_text:
                key, value = cell_text.split(':', 1)
                key = key.strip()
                value = value.strip()
                if key.lower() == 'nome':
                    nome_aluno = value
                if key.lower() == 'matricula':
                    matricula_aluno = value
            elif len(row.cells) >= 2:
                key = row.cells[0].text.strip().replace(':', '')
                value = row.cells[1].text.strip()
                if 'nome' in key.lower():
                    nome_aluno = value
                if 'matricula' in key.lower():
                    matricula_aluno = value

        # Extração do boletim, linha a linha
        idx_estado = None
        for j, row in enumerate(tabela_boletim.rows):
            linha = [cell.text.strip() for cell in row.cells]

            if j == 0:
                cabecalho = ['Nome', 'Matrícula'] + linha
                try:
                    idx_estado = linha.index('Estado')
                except ValueError:
                    idx_estado = None
                continue

            if idx_estado is not None and linha and len(linha) > idx_estado:
                estado = linha[idx_estado].upper()
                if estado != 'CANCELADO':
                    boletins_filtrados.append([nome_aluno, matricula_aluno] + linha)

    if not boletins_filtrados:
        return None, "Nenhum registro válido foi encontrado no arquivo."

    df = pd.DataFrame(boletins_filtrados, columns=cabecalho)
    return df, None


# =========================================================
# 2. Geração do Excel em memória (mesma lógica do script antigo)
# =========================================================

def gerar_excel(df_final, nota_corte, disciplinas_zeradas):
    df_export = df_final.copy()

    col_disc = 'Disciplinas abaixo da média'
    df_export['Situação'] = df_export[col_disc].apply(
        lambda x: 'Aprovado' if float(x) == 0 else ''
    )
    df_export['Observação'] = ''

    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Desempenho'

    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
    red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    headers = ['Nome'] + list(df_export.columns)

    for c, name in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=name)

    for row_num, (index, row) in enumerate(df_export.iterrows(), 2):
        ws.cell(row=row_num, column=1, value=index)
        for col_num, value in enumerate(row, 2):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            try:
                val = float(value)
                header = headers[col_num - 1]
                if header not in ['Disciplinas abaixo da média', 'Média global', 'Situação', 'Observação']:
                    if header in disciplinas_zeradas:
                        cell.fill = yellow_fill
                    elif val < nota_corte:
                        cell.fill = red_fill
                    else:
                        cell.fill = green_fill
            except:
                pass
            cell.border = border

    wb.save(output)
    output.seek(0)
    return output

# =========================================================
# 3. Geração do PDF em memória (mesma lógica do script antigo)
# =========================================================

def gerar_html_tabela(df, titulo, nota_corte, disciplinas_zeradas):
    html = f"""
    <html>
    <head>
        <style>
            @page {{ size: A4 landscape; margin: 2mm; }}
            body {{ font-family: Arial, sans-serif; font-size: 6px; }}
            table {{ width: 100%; border-collapse: collapse; }}
            th, td {{ border: 1px solid #000; padding: 1px; text-align: center; }}
            th {{ background-color: #f2f2f2; }}
            h2 {{ text-align: center; }}
            .abaixo-da-media {{ background-color: #f8d7da; color: red; }}
            .acima-da-media {{ background-color: #d4edda; color: green; }}
            .disciplina-zerada {{ background-color: #fff3cd; color: #856404; }}
            .disciplina-neutra {{ background-color: #ffffff; color: black; }}
        </style>
    </head>
    <body>
        <h2>{titulo}</h2>
        <table>
            <thead>
                <tr>
                    <th>Nome</th>
                    {"".join(f"<th>{col}</th>" for col in df.columns)}
                </tr>
            </thead>
            <tbody>
    """
    for index, row in df.iterrows():
        html += f"<tr><td>{index}</td>"
        for col_name, value in row.items():
            classe = "disciplina-neutra"
            if col_name in disciplinas_zeradas:
                classe = "disciplina-zerada"
            elif col_name != 'Disciplinas abaixo da média' and isinstance(value, (int, float)):
                classe = "abaixo-da-media" if value < nota_corte else "acima-da-media"
            html += f'<td class="{classe}">{value}</td>'
        html += "</tr>"
    html += "</tbody></table></body></html>"
    return html


def gerar_pdf(df, titulo, nota_corte, disciplinas_zeradas):
    html = gerar_html_tabela(df, titulo, nota_corte, disciplinas_zeradas)
    pdf_file = BytesIO()
    pisa.CreatePDF(io.StringIO(html), dest=pdf_file)
    pdf_file.seek(0)
    return pdf_file


def formatar_notas(val):
    try:
        return f"{float(val):.2f}"
    except (TypeError, ValueError):
        return val


def abreviar_disciplinas(df_final, num_letras):
    new_columns = []
    for col in df_final.columns:
        abreviado = ' '.join(word[:num_letras] for word in col.split())
        new_columns.append(abreviado)
    df_final.columns = new_columns
    return df_final


def highlight_cells(val, col_name, col_list, nota_corte, disciplinas_zeradas):
    try:
        val = float(val)
        if col_name == col_list[-2]:
            return 'background-color: white; color: black'
        elif col_name in disciplinas_zeradas:
            return 'background-color: #fff3cd; color: #856404'
        elif val < nota_corte:
            return 'background-color: #f8d7da; color: red'
        else:
            return 'background-color: #d4edda; color: green'
    except (TypeError, ValueError):
        return ''



# =========================================================
# 4. Interface do Streamlit
# =========================================================

st.sidebar.header("⚙️ Configurações")
nota_corte = st.sidebar.number_input("Nota de corte", min_value=0, max_value=100, value=60)
num_letras = st.sidebar.number_input("Letras para abreviação de disciplina", min_value=1, max_value=10, value=3)

arquivo_enviado = st.file_uploader(
    "Envie o boletim (apenas um arquivo .docx por vez)",
    type=["docx"],
    accept_multiple_files=False,
)

if arquivo_enviado is None:
    st.warning("Aguardando o envio de um arquivo .docx.")
    st.stop()

# O arquivo enviado já chega como objeto em memória (BytesIO).
# Não é gravado em nenhuma pasta do servidor.
df, erro = extrair_dados_docx(arquivo_enviado)

if erro:
    st.error(erro)
    st.stop()

if "Nota Final" not in df.columns:
    st.error("Coluna 'Nota Final' não encontrada no arquivo enviado.")
    st.stop()

df["Nota Final"] = pd.to_numeric(df["Nota Final"], errors="coerce").fillna(0)

if "Disciplina" not in df.columns:
    st.error("Coluna 'Disciplina' não encontrada no arquivo enviado.")
    st.stop()

# Nome da turma usado nos títulos do PDF e nos nomes dos arquivos baixados
nome_base = arquivo_enviado.name.rsplit('.', 1)[0]
titulo_turma = st.text_input("Nome da turma (usado nos arquivos gerados)", value=nome_base)

# Tabela pivot: uma linha por aluno, uma coluna por disciplina
pivot = df.pivot_table(index="Nome", columns="Disciplina", values="Nota Final", aggfunc="first").fillna(0)
disciplinas_zeradas = pivot.loc[:, (pivot == 0).all()].columns.tolist()
disciplinas_para_analisar = pivot.drop(columns=disciplinas_zeradas)

pivot['Disciplinas abaixo da média'] = (disciplinas_para_analisar < nota_corte).sum(axis=1)
pivot['Média global'] = pivot.drop(columns=['Disciplinas abaixo da média']).mean(axis=1).round(2)
pivot_sorted = pivot.sort_values(by=["Disciplinas abaixo da média", "Média global"], ascending=[False, True])

pivot_sorted_fmt = pivot_sorted.round(2).map(formatar_notas)
colunas_originais = list(pivot_sorted.columns)

st.subheader(f"📋 {titulo_turma}")
st.dataframe(
    pivot_sorted_fmt.style.apply(
        lambda col: col.apply(
            lambda val: highlight_cells(val, col.name, colunas_originais, nota_corte, disciplinas_zeradas)
        ),
        axis=0,
    ),
    use_container_width=True,
)

# Versão com nomes de disciplina abreviados
pivot_abreviado = abreviar_disciplinas(pivot_sorted.copy(), num_letras)
disciplinas_zeradas_abreviadas = pivot_abreviado.loc[:, (pivot_abreviado == 0).all()].columns.tolist()
pivot_abreviado_fmt = pivot_abreviado.map(formatar_notas)
colunas_abreviadas = list(pivot_abreviado.columns)

st.subheader(f"📋 Disciplinas abreviadas ({num_letras} letras)")
st.dataframe(
    pivot_abreviado_fmt.style.apply(
        lambda col: col.apply(
            lambda val: highlight_cells(val, col.name, colunas_abreviadas, nota_corte, disciplinas_zeradas_abreviadas)
        ),
        axis=0,
    ),
    use_container_width=True,
)

# =========================================================
# 5. Downloads (Excel e PDF gerados só em memória)
# =========================================================

excel_completo = gerar_excel(pivot_sorted, nota_corte, disciplinas_zeradas)
pdf_bytes = gerar_pdf(pivot_sorted, titulo_turma, nota_corte, disciplinas_zeradas)

col1, col2 = st.columns(2)
with col1:
    st.download_button(
        label="📥 Baixar Excel",
        data=excel_completo,
        file_name=f"{nome_base}_completo.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
with col2:
    st.download_button(
        label="📄 Baixar PDF",
        data=pdf_bytes,
        file_name=f"{nome_base}.pdf",
        mime="application/pdf",
    )
